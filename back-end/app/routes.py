from flask import request, Response, current_app
from flask_restx import Resource, Api
from openpyxl import load_workbook
from werkzeug.utils import secure_filename
from google.oauth2 import id_token
from google.auth.transport import requests as google_request
import requests
import pandas as pd
import numpy as np
import json
import os
import re
import random

from app.models import create_data_model
from app.models import create_account_model
from app.utils import assign_group
from app.utils import add_code_to_callRule
from app.utils import calculate_amortization_schedule

GOOGLE_CLIENT_ID = "864707476707-u61028sfa4isftnerqaralk9df0tta05.apps.googleusercontent.com"

def register_routes(api: Api):
    data_model = create_data_model(api)
    account_model = create_account_model(api)

    @api.route('/get_all_houses')
    class GetAllHouse(Resource):
        def get(self):
            file_path = 'D:\\DATN\\model\\houseAll.xls'
            
            if not os.path.exists(file_path):
                return Response(json.dumps({"error": "File not found"}), content_type="application/json", status=404)
            
            df = pd.read_excel(file_path)
            df = df.replace({np.nan: None})
            data = df.to_dict(orient='records')
            data_json = json.dumps(data, ensure_ascii=False, indent=4)
            return Response(data_json, content_type="application/json; charset=utf-8")
    
    @api.route('/get_house/<meta_code>')
    class GetHouseByMetaCode(Resource):
        def get(self, meta_code):
            file_path = 'D:\\DATN\\model\\houseAll.xls'
            
            if not os.path.exists(file_path):
                return Response(json.dumps({"error": "File not found"}), content_type="application/json", status=404)
            
            df = pd.read_excel(file_path)
            df = df.replace({np.nan: None})
            house = df[df['meta_code'] == meta_code].to_dict(orient='records')
            
            if not house:
                return Response(json.dumps({"error": "House not found"}), content_type="application/json", status=404)
            
            data_json = json.dumps(house[0], ensure_ascii=False, indent=4)
            return Response(data_json, content_type="application/json; charset=utf-8")
        
    @api.route('/get_houses_from_model/<id>')
    class GetHousesFromModel(Resource):
        def get(self, id):
            file_path = 'D:\\DATN\\model\\data.xlsx'
            notebook_path = 'D:\\DATN\\model\\callRule.ipynb'
            wb = load_workbook(filename=file_path)
            buyers_sheet = wb['buyers']
            row_number = None
            
            for row in buyers_sheet.iter_rows(min_row=2, max_row=buyers_sheet.max_row):
                if row[0].value == id:
                    row_number = row[0].row - 2
                    break
            
            if row_number is None:
                return Response(json.dumps({"error": "ID not found in the buyers sheet"}), content_type="application/json", status=404)

            add_code_to_callRule(notebook_path, id, row_number)
            
            house_path = f'D:\\DATN\\model\\testOutput\\cust_{id}.xlsx'
            
            if not os.path.exists(house_path):
                return Response(json.dumps({"error": "File not found"}), content_type="application/json", status=404)
            
            df = pd.read_excel(house_path)
            df = df.drop(df.columns[[1, 2, 6]], axis=1)
            df = df.replace({np.nan: None})
            
            df = df[df['housePerformance'] >= 70]
            df = df.sort_values(by='housePerformance', ascending=False)
            
            data = df.to_dict(orient='records')
            data_json = json.dumps(data, ensure_ascii=False, indent=4)
            return Response(data_json, content_type="application/json; charset=utf-8", status=200)   
        
    @api.route('/amortization-schedule')
    class AmortizationSchedule(Resource):
        def post(self):
            request_data = request.get_json()
            
            loan_amount = request_data.get('loanAmount')
            annual_interest_rate = request_data.get('annualInterestRate')
            loan_term_in_months = request_data.get('loanTermInMonths')

            if not loan_amount or not annual_interest_rate or not loan_term_in_months:
                return 'Missing required parameters', 400

            try:
                loan_amount = float(loan_amount)
                annual_interest_rate = float(annual_interest_rate)
                loan_term_in_months = int(loan_term_in_months)
            except ValueError:
                return 'Invalid parameters, must be numbers', 400

            # Tính toán lịch trả nợ
            amortization_schedule = calculate_amortization_schedule(loan_amount, annual_interest_rate, loan_term_in_months)

            total_interest = sum(item['interest'] for item in amortization_schedule)
            max_payment_info = max(amortization_schedule, key=lambda x: x['total_payment'])
            total_payment = sum(item['total_payment'] for item in amortization_schedule)
            interest_to_payment_ratio = total_interest / total_payment
            
            return {
                'Total Interest': total_interest,
                'Max Monthly Payment': max_payment_info['total_payment'],
                'Max Payment Month': max_payment_info['month'],
                'Total Payment': total_payment,
                'Interest to Payment Ratio': interest_to_payment_ratio,
                'Schedule': amortization_schedule
            }, 200
        
    @api.route('/google_login')
    class GoogleLogin(Resource):
        def post(self):
            file_path = 'D:\\DATN\\model\\data.xlsx'
            wb = load_workbook(filename=file_path)
            buyers_info_sheet = wb['buyers_info']
            buyers_sheet = wb['buyers']
            
            sheet_data = pd.DataFrame(buyers_info_sheet.values)
            
            columns = sheet_data.iloc[0]
            df = sheet_data[1:]
            df.columns = columns
            
            data = request.get_json()
            token = data.get('token')
            if not token:
                return 'Token is missing', 400
            
            try:
                idinfo = id_token.verify_oauth2_token(token, google_request.Request(), GOOGLE_CLIENT_ID)
                
                email = idinfo['email']
                name = idinfo['name']
                image = idinfo.get('picture', '')
                
                if email not in df['email'].values:     
                    if image:
                        image_response = requests.get(image)
                        folder_path = current_app.config['UPLOAD_IMAGE']
                        if image_response.status_code == 200:
                            image_path = os.path.join(folder_path, f"{image.replace('https://lh3.googleusercontent.com/a/', '')}.jpg")
                            with open(image_path, 'wb') as file:
                                file.write(image_response.content)
                        else:
                            image_path = ''
                    else:
                        image_path = ''

                    header = [cell.value for cell in buyers_info_sheet[1]]
                    id_col = header.index('id')
                    name_col = header.index('name')
                    email_col = header.index('email')
                    image_col = header.index('image')

                    header_buyer = [cell.value for cell in buyers_sheet[1]]
                    id_buyer_col = header.index('id')

                    buyers_info_none_rows = [row for row in buyers_info_sheet.iter_rows() if all(cell.value is None for cell in row)]
                    for row in buyers_info_none_rows:
                        buyers_info_sheet.delete_rows(row[0].row)

                    buyers_none_rows = [row for row in buyers_sheet.iter_rows() if all(cell.value is None for cell in row)]
                    for row in buyers_none_rows:
                        buyers_sheet.delete_rows(row[0].row)

                    new_row = [None] * len(header)
                     # get id
                    last_row_id = None
                    for row in buyers_info_sheet.iter_rows(min_row=2, max_col=1, max_row=buyers_info_sheet.max_row):
                        cell_value = row[0].value
                        if cell_value:
                            last_row_id = cell_value

                    if last_row_id:
                        prefix, numeric_part = re.match(r'([^\d]+)(\d+)', last_row_id).groups()
                        next_numeric_part = int(numeric_part) + 1
                        new_row[id_col] = f'{prefix}{next_numeric_part}'
                    else:
                        new_row[id_col] = 'x1'

                    new_row[name_col] = name
                    new_row[email_col] = email
                    new_row[image_col] = f"{image.replace('https://lh3.googleusercontent.com/a/', '')}.jpg"

                    buyers_info_sheet.append(new_row)

                    new_row_buyer = [None] * len(header_buyer)
                    new_row_buyer[id_buyer_col] = new_row[id_col]
                    buyers_sheet.append(new_row_buyer)
                    wb.save(filename=file_path)
                    
                    new_user_info = {
                        "id": new_row[id_col],
                        "name": name,
                        "email": email,
                        "image": new_row[image_col]
                    }
                    return new_user_info, 200   
                else:
                    user = df[(df['email'] == email)]
                    
                    if not user.empty:
                        user_info = user.iloc[0].to_dict()
                        return user_info, 200
                    else:
                        return 'Không tìm thấy người dùng', 401
                
            except ValueError:
                return 'Invalid token', 400
    
    @api.route('/login')
    class Login(Resource):
        @api.expect(account_model)
        def post(self):
            file_path = 'D:\\DATN\\model\\data.xlsx'
            wb = load_workbook(filename=file_path)
            buyers_info_sheet = wb['buyers_info']
            
            data = pd.DataFrame(buyers_info_sheet.values)
            
            columns = data.iloc[0]
            df = data[1:]
            df.columns = columns
            
            user_data = request.get_json()
            
            emailOrPhone = user_data['emailOrPhone']
            password = user_data['password']

            # Check if the user input is an email or a phone number
            if '@' in emailOrPhone:  # If '@' is present, treat it as an email
                match_column = 'email'
            else:  # Otherwise, treat it as a phone number
                match_column = 'phone_number'

            # Check if emailOrPhone exists in the data
            if emailOrPhone not in df[match_column].values:
                return 'Email hoặc số điện thoại không tồn tại', 400

            # Check if the password matches the emailOrPhone
            user = df[(df[match_column] == emailOrPhone) & (df['password'] == password)]

            if not user.empty:
                user_info = user.iloc[0].to_dict()
                return user_info, 200
            else:
                return 'Mật khẩu không đúng', 401
            
    @api.route('/register')
    class Register(Resource):
        @api.expect(account_model)
        def post(self):
            file_path = 'D:\\DATN\\model\\data.xlsx'
            wb = load_workbook(filename=file_path)
            buyers_info_sheet = wb['buyers_info']
            buyers_sheet = wb['buyers']
            
            data = request.get_json()   
            
            emailOrPhone = data.get('emailOrPhone')
            password = data.get('password')
            
            header = [cell.value for cell in buyers_info_sheet[1]]
            id_col = header.index('id')
            name_col = header.index('name')
            email_col = header.index('email')
            phone_col = header.index('phone_number')
            password_col = header.index('password')
            
            header_buyer_sheet = [cell.value for cell in buyers_sheet[1]]
            id_buyer_col = header_buyer_sheet.index('id')
            
            for row in buyers_info_sheet.iter_rows(min_row=2, values_only=True):
                if (row[email_col] == emailOrPhone) or (row[phone_col] == emailOrPhone):
                    return 'Email hoặc số điện thoại đã tồn tại', 400
                
            buyers_info_none_rows = [row for row in buyers_info_sheet.iter_rows() if all(cell.value is None for cell in row)]
            for row in buyers_info_none_rows:
                buyers_info_sheet.delete_rows(row[0].row)
                
            buyers_none_rows = [row for row in buyers_sheet.iter_rows() if all(cell.value is None for cell in row)]
            for row in buyers_none_rows:
                buyers_sheet.delete_rows(row[0].row)
                
            new_row = [None] * len(header)
             # get id
            last_row_id = None
            for row in buyers_info_sheet.iter_rows(min_row=2, max_col=1, max_row=buyers_info_sheet.max_row):
                cell_value = row[0].value
                if cell_value:
                    last_row_id = cell_value
                    
            if last_row_id:
                prefix, numeric_part = re.match(r'([^\d]+)(\d+)', last_row_id).groups()
                next_numeric_part = int(numeric_part) + 1
                new_row[id_col] = f'{prefix}{next_numeric_part}'
            else:
                new_row[id_col] = 'x1'
                
            random_username = f'user{random.randint(1000, 9999)}'
            new_row[name_col] = random_username
                
            if '@' in emailOrPhone:
                new_row[email_col] = emailOrPhone
            else:
                new_row[phone_col] = emailOrPhone
            new_row[password_col] = password
            
            buyers_info_sheet.append(new_row)
            
            new_row_buyer = [None] * len(header_buyer_sheet)
            new_row_buyer[id_buyer_col] = new_row[id_col]
            buyers_sheet.append(new_row_buyer)
            wb.save(filename=file_path)
            
            return 'Created account successfully', 200

        @api.route('/get_user/<id>')
        class GetUser(Resource):
            def get(self, id):
                file_path = 'D:\\DATN\\model\\data.xlsx'
                wb = load_workbook(filename=file_path)
                buyers_sheet = wb['buyers']
                buyers_info_sheet = wb['buyers_info']
        
                user_info = {}
        
                # Search for user in buyers_info_sheet
                for row in buyers_info_sheet.iter_rows(min_row=2, max_row=buyers_info_sheet.max_row):
                    if row[0].value == id:
                        user_info.update({
                            "id": row[0].value,
                            "name": row[1].value,
                            "email": row[2].value,
                            "password": row[3].value,
                            "phone_number": row[4].value,
                            "address": row[5].value,
                            "image": row[6].value
                            # Add more fields as needed
                        })
                        break
                    
                # Search for user in buyers_sheet
                for row in buyers_sheet.iter_rows(min_row=2, max_row=buyers_sheet.max_row):
                    if row[0].value == id:
                        user_info.update({
                            "group": row[1].value,
                            "age": row[2].value,
                            "is_married": row[3].value,
                            "no_child": row[4].value,
                            "month_income": row[5].value,
                            "monthly_amt": row[6].value,
                            "avai_amt": row[7].value,
                            "desired_location": row[8].value,
                            "desired_interiorStatus": row[9].value
                            # Add more fields as needed
                        })
                        break
                    
                if user_info:
                    json_data = json.dumps(user_info, ensure_ascii=False, indent=4)
                    return Response(json_data, content_type="application/json; charset=utf-8")
        
                return {"error": "User not found"}, 400
      
    @api.route('/update_user/<id>')
    class UpdateUser(Resource):
        @api.expect(data_model)
        def put(self, id):
            data = request.get_json()
            formData = data.get('formData', {})
            
            file_path = 'D:\\DATN\\model\\data.xlsx'
            wb = load_workbook(filename=file_path)
            buyers_sheet = wb['buyers']
            buyers_info_sheet = wb['buyers_info']
            
            header = [cell.value for cell in buyers_info_sheet[1]]
            email_col = header.index('email')
            phone_col = header.index('phone_number')
            
            for row in buyers_info_sheet.iter_rows(min_row=2, max_row=buyers_info_sheet.max_row):
                if row[0].value != id:
                    if row[email_col].value == formData.get('email'):
                        return 'Email đã tồn tại', 400
                    elif row[phone_col].value == formData.get('phone_number'):
                        return 'Số điện thoại đã tồn tại', 400
            
            # Find and update the row in 'buyers_info' sheet
            for row in buyers_info_sheet.iter_rows(min_row=2, max_row=buyers_info_sheet.max_row):
                if row[0].value == id:
                    row[1].value = formData.get('name', row[1].value)
                    row[2].value = formData.get('email', row[2].value)
                    row[4].value = formData.get('phone_number', row[4].value)
                    row[5].value = formData.get('address', row[5].value)
                    row[6].value = formData.get('image', row[6].value)
                    break
            
            # Find and update the row in 'buyers' sheet
            for row in buyers_sheet.iter_rows(min_row=2, max_row=buyers_sheet.max_row):
                if row[0].value == id:
                    row[1].value = assign_group(formData)
                    row[2].value = int(formData.get('age', row[2].value))
                    row[3].value = formData.get('is_married', row[3].value)
                    row[4].value = formData.get('no_child', row[4].value)
                    row[5].value = int(formData.get('month_income', str(row[5].value)).replace(',', ''))
                    row[6].value = int(formData.get('monthly_amt', str(row[6].value)).replace(',', ''))
                    row[7].value = int(formData.get('avai_amt', str(row[7].value)).replace(',', ''))
                    row[8].value = formData.get('desired_location', row[8].value)
                    row[9].value = formData.get('desired_interiorStatus', row[9].value)
                    break
                
            wb.save(filename=file_path)
            
            return 'User updated successfully!', 200
        
    @api.route('/upload_image')
    class UploadImage(Resource):
        def post(self):
            file = request.files['file']
            
            if file:
                filename = secure_filename(file.filename)
                upload_folder = current_app.config['UPLOAD_IMAGE']
                file_path = os.path.join(upload_folder, filename)
                file.save(file_path)
                return Response(json.dumps({"succuess": "File upload successfully", "file_path": file_path}), content_type="application/json", status=200)
            
    @api.route('/change_password/<id>')
    class ChangePassword(Resource):
        @api.expect(data_model)
        def put(self, id):
            data = request.get_json()
            
            file_path = 'D:\\DATN\\model\\data.xlsx'
            wb = load_workbook(filename=file_path)
            buyers_info_sheet = wb['buyers_info']
            
            # Find and update the row in 'buyers_info' sheet
            for row in buyers_info_sheet.iter_rows(min_row=2, max_row=buyers_info_sheet.max_row):
                if row[0].value == id:
                    row[3].value = data.get('password', row[3].value)
                    break
                
            wb.save(filename=file_path)
            
            return 'Changed password successfully!', 200
        
    @api.route('/forgot_password')
    class ForgotPassword(Resource):
        # @api.expect(data_model)
        def put(self):
            file_path = 'D:\\DATN\\model\\data.xlsx'
            wb = load_workbook(filename=file_path)
            buyers_info_sheet = wb['buyers_info']
            
            data = request.get_json()
            
            emailOrPhone = data.get('emailOrPhone')
            new_password = data.get('newPassword')
            
            match_column = 2 if '@' in emailOrPhone else 4
            
            user_row = None
            for row in buyers_info_sheet.iter_rows(min_row=2, max_row=buyers_info_sheet.max_row):
                if row[match_column].value == emailOrPhone:
                    user_row = row
                    break
                
            if not user_row:
                return 'Email hoặc số điện thoại không tồn tại', 400
            else:
                user_row[3].value = new_password
                wb.save(filename=file_path)
                return 'Password updated successfully', 200
            
            